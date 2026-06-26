"""F02 (.xlsm) 解析器：官方表單 → F02Form。

讀取分工：
  - openpyxl（data_only=True）：評估表佈局不規則且含公式（I2 分數、I4 分級、E46:H46 百分比），
    用儲存格定址最清楚；同時拿到 Excel 已算好的「快取分數」作為比對基準。
  - pandas：剩餘風險評鑑表是規則表格，用 DataFrame 取「剩餘風險評鑑分數」欄最大值最乾淨。
"""

from __future__ import annotations

import warnings
from pathlib import Path

import openpyxl
import pandas as pd

from govcheck.models import CachedScores, DomainScores, F02Form
from govcheck.scoring.f02_score import load_config

ASSESS_SHEET = "AI系統固有風險分級評估表"
RESIDUAL_SHEET = "AI系統剩餘風險評鑑表"
TREATMENT_SHEET = "AI系統風險處理計畫表"

# 評估表中四個風險域百分比所在欄（row 46）與分數/分級格
PCT_CELLS = {"finance": "E46", "operation": "F46", "reputation": "G46", "compliance": "H46"}
OVERALL_CELL = "I2"
GRADE_CELL = "I4"
UPLIFT_ROWS = {"factor1": 42, "factor2": 43, "factor3": 44}
RESIDUAL_SCORE_COL = "剩餘風險 評鑑分數"


def parse_f02(path: str | Path, cfg: dict | None = None) -> F02Form:
    cfg = cfg or load_config()
    path = Path(path)
    question_ids = set(cfg["questions"].keys())

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    ws = wb[ASSESS_SHEET]

    answers: dict[str, str | None] = {}
    for row in ws.iter_rows():
        qid_cell = row[0].value  # A 欄題號
        if isinstance(qid_cell, str) and qid_cell.strip() in question_ids:
            qid = qid_cell.strip()
            answers[qid] = _norm_answer(row[2].value)  # C 欄答案

    uplift = {name: _norm_answer(ws[f"C{r}"].value) for name, r in UPLIFT_ROWS.items()}

    # 快取百分比：若四格皆空（如 openpyxl 產生、未經 Excel 計算的檔），視為「無快取」None，
    # 讓計分比對規則略過，而非誤判為全 0。
    raw_pct = {d: _num_or_none(ws[c].value) for d, c in PCT_CELLS.items()}
    has_pct = any(v is not None for v in raw_pct.values())
    cached = CachedScores(
        percentages=DomainScores(**{d: (v or 0.0) for d, v in raw_pct.items()}) if has_pct else None,
        overall=_num_or_none(ws[OVERALL_CELL].value),
        grade=_clean(ws[GRADE_CELL].value),
    )

    residual_filled, residual_max = _read_residual(path)
    treatment_filled = _sheet_has_data(wb, TREATMENT_SHEET)
    wb.close()

    return F02Form(
        answers=answers,
        uplift=uplift,
        cached=cached,
        residual_filled=residual_filled,
        residual_max_score=residual_max,
        treatment_filled=treatment_filled,
    )


def _read_residual(path: Path) -> tuple[bool, float | None]:
    """用 pandas 讀剩餘風險評鑑表：是否有資料列、剩餘風險分數欄最大值。"""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = pd.read_excel(path, sheet_name=RESIDUAL_SHEET, engine="openpyxl")
    df = df.dropna(how="all")
    if df.empty:
        return False, None
    score_col = next((c for c in df.columns if str(c).replace("\n", " ").strip() == RESIDUAL_SCORE_COL), None)
    max_score = None
    if score_col is not None:
        nums = pd.to_numeric(df[score_col], errors="coerce").dropna()
        if not nums.empty:
            max_score = float(nums.max())
    return True, max_score


def _sheet_has_data(wb, sheet_name: str) -> bool:
    """表頭以下是否有任何非空資料列。"""
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳過表頭
        if any(v is not None and str(v).strip() for v in row):
            return True
    return False


def _norm_answer(v) -> str | None:
    s = _clean(v)
    if not s:
        return None
    s = s.strip().upper().replace("是", "Y").replace("否", "N")
    return s if s in {"Y", "N"} else None


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num_or_none(v) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None
