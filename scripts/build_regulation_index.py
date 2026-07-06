"""Build-time RAG 管線：法規索引（build ①③）+ canonical 檢索映射（build ②④）。

seam 函式（Lead 鎖定命名）：
  build ①  build_index()           — PDF→chunk→embed→Milvus Lite + index_meta.json
  build ③  print_chunk_tree()      — --dry-run：印 chunk 樹，不 embed 不寫庫
  build ②  build_retrieval_map()   — F03/F02 canonical 抽取 → 預算 → retrieval_map.json（原子寫）
  build ④  run_eval()              — recall@k + 分數分佈報告（stdout）
  CLI 入口  main()                  — argparse 統一分派

T1 模組（pdf_text / chunker / refs）的 import 一律放在函式體內（lazy），
使腳本在 T1 落地前亦可測試結構；T2 模組（config / embedding / store）為 top-level import。

注意：
- --dry-run 只印 stdout；不得寫出任何受 git 追蹤路徑下的檔案（AC-17）。
- stdout 報告可含統計與 section_path；程式碼中以 NOTE 註記「輸出不得貼進會 commit 的檔案」。
- build 產物（Milvus .db / index_meta.json / retrieval_map.json）落 data/（gitignored）。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

from govcheck.logging_setup import get_logger
from govcheck.parsers.f02_parser import ASSESS_SHEET
from govcheck.rag.config import load_rag_config
from govcheck.rag.embedding import EmbeddingClient, EmbeddingError
from govcheck.rag.store import RegulationStore

log = get_logger("build_regulation_index")

ROOT = Path(__file__).resolve().parents[1]

# ── 法規冊登錄（reg_code → chunker style）────────────────────────────────────
_REG_REGISTRY: list[dict] = [
    {"reg_code": "R01", "style": "article"},
    {"reg_code": "R02", "style": "chapter"},
    {"reg_code": "R03", "style": "chapter"},
    {"reg_code": "R04", "style": "chapter"},
    {"reg_code": "R05", "style": "chapter"},
    {"reg_code": "R06", "style": "chapter"},
    {"reg_code": "R07", "style": "chapter"},
]

_EMBED_BATCH = 32  # 每批 embed 筆數（spec §4.6 ①）


# ── 內部工具 ─────────────────────────────────────────────────────────────────

def _find_pdf(reg_code: str) -> Path | None:
    data_dir = ROOT / "data" / "original"
    for pattern in (f"{reg_code}*.pdf", f"{reg_code}*.PDF"):
        matches = list(data_dir.glob(pattern))
        if matches:
            return matches[0]
    return None


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _check_gitignore(path_str: str) -> bool:
    """Return True if path_str is listed as ignored by git check-ignore."""
    try:
        result = subprocess.run(
            ["git", "check-ignore", path_str],
            capture_output=True,
            cwd=str(ROOT),
            timeout=5,
        )
        return result.returncode == 0
    except (subprocess.SubprocessError, FileNotFoundError):
        return False


def _read_f03_items() -> list[dict]:
    """從 data/original 官方附件三讀 20 個 F03 檢核項（item_id/topic/description/canonical_ref_raw）。

    NOTE: 輸出不得貼進會 commit 的檔案。
    """
    import re
    import warnings

    import openpyxl

    from govcheck.parsers._util import clean
    from govcheck.review.config import load_review_config

    rcfg = load_review_config()
    f3 = rcfg["f03"]
    chk = f3["checklist"]

    candidates = list((ROOT / "data" / "original").glob("附件三*"))
    if not candidates:
        raise FileNotFoundError(f"附件三 not found in {ROOT / 'data' / 'original'}")
    f03_path = candidates[0]

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(f03_path, data_only=True)

    ws = wb[f3["sheet"]]
    id_col = chk["item_id_col"]
    topic_col = chk["topic_col"]
    desc_col = chk["item_desc_col"]
    ref_col = chk.get("regulation_ref_col", "L")  # L 欄=規範參考（T4 補入 config 後讀取；暫 hardcode L）
    start_row: int = chk["data_start_row"]
    pattern = re.compile(chk["item_id_pattern"])

    items: list[dict] = []
    for row_idx in range(start_row, start_row + 30):  # 最多 30 列（含備註）
        item_id_val = clean(ws[f"{id_col}{row_idx}"].value)
        if not item_id_val or not pattern.match(item_id_val):
            continue
        items.append(
            {
                "item_id": item_id_val,
                "topic": clean(ws[f"{topic_col}{row_idx}"].value),
                "description": clean(ws[f"{desc_col}{row_idx}"].value),
                "canonical_ref_raw": clean(ws[f"{ref_col}{row_idx}"].value),
            }
        )
    wb.close()
    log.info("read f03 items n=%d", len(items))
    return items


def _read_f02_questions(f02_path: Path | None = None) -> dict[str, dict]:
    """從 data/original 官方附件二讀 F02 題文（qid → question_text）。

    僅 build 時記憶體使用；絕不寫入任何會 commit 的路徑（spec §4.2 / 隱私 H1）。
    NOTE: 輸出不得貼進會 commit 的檔案。
    """
    import warnings

    import openpyxl

    if f02_path is None:
        candidates = list((ROOT / "data" / "original").glob("附件二*"))
        if not candidates:
            raise FileNotFoundError(f"附件二 not found in {ROOT / 'data' / 'original'}")
        f02_path = candidates[0]

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(f02_path, data_only=True)

    # 使用 parser 常數，避免與 f02_parser.ASSESS_SHEET 漂移（issue 4）
    ws = wb[ASSESS_SHEET]
    raw: dict[str, dict] = {}
    for row in range(2, 47):
        qid_val = ws[f"A{row}"].value
        text_val = ws[f"B{row}"].value
        if isinstance(qid_val, str) and qid_val.strip():
            qid = qid_val.strip()
            raw[qid] = {
                "question_text": (str(text_val).strip() if text_val else "")
            }
    wb.close()

    # 以 scoring config 的 qid 集合為權威，交集過濾：
    # sheet 只提供題文供 embedding，scored qid 集合由 f02_score.load_config() 決定（issue 3）
    from govcheck.scoring.f02_score import load_config as load_scoring_config
    scored_qids: set[str] = set(load_scoring_config()["questions"].keys())
    questions = {qid: v for qid, v in raw.items() if qid in scored_qids}

    log.info(
        "read f02 questions raw=%d scored=%d filtered=%d",
        len(raw),
        len(scored_qids),
        len(questions),
    )
    return questions


# ── seam 函式 ────────────────────────────────────────────────────────────────


def iter_regulation_chunks(cfg: dict | None = None):
    """逐一產出所有 7 份法規的 RegulationChunk（generator）。

    lazy import T1 模組（pdf_text / chunker）；data/original 缺 PDF 則 skip 並 WARNING。
    供 build ① 與 --dry-run 消費。
    NOTE: 輸出不得貼進會 commit 的檔案（chunk.text 含法規摘錄）。
    """
    from govcheck.rag.chunker import chunk_regulation
    from govcheck.rag.pdf_text import load_clean_lines

    for reg_info in _REG_REGISTRY:
        reg_code = reg_info["reg_code"]
        pdf_path = _find_pdf(reg_code)
        if pdf_path is None:
            log.warning("reg PDF not found reg=%s", reg_code)
            continue
        reg_title = pdf_path.stem
        log.debug("chunking reg=%s path=%s", reg_code, pdf_path.name)
        lines = load_clean_lines(pdf_path)
        chunks = chunk_regulation(
            lines,
            reg_code=reg_code,
            reg_title=reg_title,
            style=reg_info["style"],
        )
        log.debug("reg=%s chunks=%d", reg_code, len(chunks))
        yield from chunks


def print_chunk_tree(cfg: dict | None = None) -> None:
    """--dry-run ③：印 chunk 樹（reg → section_path → title/字數）到 stdout；不 embed 不寫庫。

    AC-17：僅印 stdout；不寫出任何受 git 追蹤路徑下的檔案。
    NOTE: 輸出不得貼進會 commit 的檔案。
    """
    from collections import defaultdict

    reg_chunks: dict[str, list] = defaultdict(list)
    for chunk in iter_regulation_chunks(cfg):
        reg_chunks[chunk.reg_code].append(chunk)

    total = 0
    for reg_code, chunks in reg_chunks.items():
        print(f"\n{reg_code} ({chunks[0].reg_title}): {len(chunks)} chunks")
        for chunk in chunks:
            body_len = len(chunk.text) - len(chunk.text.split("\n")[0])  # 扣 breadcrumb
            print(f"  [{chunk.order:3d}] {chunk.section_path}  {chunk.title!r}  ({body_len} chars)")
        total += len(chunks)
    print(f"\n=== total chunks: {total} ===")


def build_index(
    cfg: dict | None = None,
    store: RegulationStore | None = None,
    embedding_client: EmbeddingClient | None = None,
) -> None:
    """build ①：PDF → chunk → embed → Milvus Lite + index_meta.json。

    每批 _EMBED_BATCH 筆呼叫 EmbeddingClient.embed；失敗時 EmbeddingError 向上傳播。
    """
    from govcheck.rag.models import IndexMeta

    cfg = cfg or load_rag_config()
    dim: int = cfg["embedding_dim"]

    _owns_store = store is None
    if store is None:
        store = RegulationStore.from_config(cfg)
    if embedding_client is None:
        embedding_client = EmbeddingClient.from_config(cfg)

    store.recreate(dim)

    all_chunks = list(iter_regulation_chunks(cfg))
    chunks_per_reg: dict[str, int] = {}
    for chunk in all_chunks:
        chunks_per_reg[chunk.reg_code] = chunks_per_reg.get(chunk.reg_code, 0) + 1

    # Batch embed + insert
    total_inserted = 0
    for batch_start in range(0, len(all_chunks), _EMBED_BATCH):
        batch = all_chunks[batch_start : batch_start + _EMBED_BATCH]
        texts = [chunk.text for chunk in batch]
        vectors = embedding_client.embed(texts)
        store.insert(batch, vectors)
        total_inserted += len(batch)
        log.debug("build_index progress=%d/%d", total_inserted, len(all_chunks))

    # sidecar
    source_sha256: dict[str, str] = {}
    for reg_info in _REG_REGISTRY:
        pdf_path = _find_pdf(reg_info["reg_code"])
        if pdf_path:
            source_sha256[pdf_path.name] = _sha256_file(pdf_path)

    meta = IndexMeta(
        schema_version=1,
        built_at=datetime.now(tz=timezone.utc).isoformat(),
        embedding_model=cfg["embedding_model"],
        embedding_dim=dim,
        chunks_per_reg=chunks_per_reg,
        source_sha256=source_sha256,
    )
    store.write_meta(meta)
    log.info(
        "build_index done regs=%d total_chunks=%d",
        len(chunks_per_reg),
        total_inserted,
    )

    if _owns_store:
        store.close()


def build_retrieval_map(
    *,
    cfg: dict | None = None,
    store=None,
    embedding_client: EmbeddingClient | None = None,
    items: list | None = None,
    f02_questions: dict | None = None,
    output_path: str | Path | None = None,
) -> None:
    """build ②：從官方附件三/二抽 canonical → curated+semantic 預算 → retrieval_map.json（原子寫）。

    原子寫：先寫 .tmp，再 os.replace，避免半寫檔被 runtime 讀到。
    gitignore 防護：未指定 output_path 時，自動執行 git check-ignore 確認產物不入庫；
    未通過即中止（data/rag/ 應已在 .gitignore）。

    可注入 store / embedding_client / items / f02_questions / output_path，供測試用。
    """
    from govcheck.rag.refs import filter_chunks_by_ref, parse_regulation_refs

    cfg = cfg or load_rag_config()

    # Determine output path + gitignore check
    # 一律錨定 ROOT，確保閘門檢查與實際寫入路徑一致（issue 2：CWD 依賴修正）
    use_default = output_path is None
    if use_default:
        output_path = ROOT / cfg["mapping_path"]
        output_path_str = str(output_path)
        if not _check_gitignore(output_path_str):
            raise RuntimeError(
                f"Abort: `git check-ignore {output_path_str}` 未通過。"
                " 請確認 .gitignore 已排除 data/rag/，避免含法規摘錄的產物誤入庫。"
            )
    else:
        output_path = Path(output_path)

    # Store
    _owns_store = store is None
    if store is None:
        store = RegulationStore.from_config(cfg)
        store.open()

    # Embedding client
    if embedding_client is None:
        embedding_client = EmbeddingClient.from_config(cfg)

    # Items & F02 questions
    if items is None:
        items = _read_f03_items()
    if f02_questions is None:
        f02_questions = _read_f02_questions()

    max_cap: int = cfg["max_sections_per_item"]
    max_excerpt: int = cfg["max_excerpt_chars"]
    top_k: int = cfg["top_k"]
    score_threshold: float | None = cfg.get("score_threshold")

    f03_items_out: dict = {}
    f02_questions_out: dict = {}

    # ── Process F03 items ────────────────────────────────────────────────────
    for item in items:
        item_id = item["item_id"]
        topic = item.get("topic") or ""
        description = item.get("description") or ""
        canonical_ref_raw = item.get("canonical_ref_raw")

        curated_sections: list[dict] = []
        refs_out: list[dict] = []

        # curated path（L 欄 → parse → lookup → filter_chunks_by_ref per-ref cap）
        if canonical_ref_raw:
            try:
                parsed_refs = parse_regulation_refs(canonical_ref_raw)
            except Exception as exc:
                log.warning("refs parse failed item=%s err=%s", item_id, type(exc).__name__)
                parsed_refs = []

            for ref in parsed_refs:
                rows = store.lookup(ref.reg_code, ref.section_path_prefix)
                matched = filter_chunks_by_ref(ref, rows)  # per-ref cap=3（refs.PER_REF_CAP）
                for row in matched:
                    curated_sections.append(
                        {
                            "reg_code": row["reg_code"],
                            "section_path": row["section_path"],
                            "title": row["title"],
                            "excerpt": row["text"][:max_excerpt],
                            "score": None,
                            "origin": "curated",
                        }
                    )
                refs_out.append(
                    {
                        "reg_code": ref.reg_code,
                        "section_path_prefix": ref.section_path_prefix,
                    }
                )

        # semantic path（embed topic:description → store.search → threshold）
        query_text = f"{topic}:{description}"
        try:
            vectors = embedding_client.embed([query_text])
            hits = store.search(vectors[0], top_k)
        except (EmbeddingError, Exception) as exc:
            log.warning("semantic search failed item=%s err=%s", item_id, type(exc).__name__)
            hits = []

        # merge：curated 優先；semantic 與任一 curated 同 (reg_code, section_path)
        # 或互為 prefix 者剔除（spec p3-02 §4.2 curated 優先、semantic 補充去重）
        curated_keys: set[tuple] = {(s["reg_code"], s["section_path"]) for s in curated_sections}
        semantic_sections: list[dict] = []

        for hit in hits:
            if score_threshold is not None and (hit.get("score") or 0.0) < score_threshold:
                continue
            hr, hp = hit["reg_code"], hit["section_path"]
            # dedup check：同 key 或互為 prefix（以 "/" 分段）
            duplicate = False
            for cr, cp in curated_keys:
                if cr != hr:
                    continue
                if cp == hp or hp.startswith(cp + "/") or cp.startswith(hp + "/"):
                    duplicate = True
                    break
            if duplicate:
                continue
            semantic_sections.append(
                {
                    "reg_code": hr,
                    "section_path": hp,
                    "title": hit["title"],
                    "excerpt": hit["text"][:max_excerpt],
                    "score": hit.get("score"),
                    "origin": "semantic",
                }
            )

        # 合併並 cap
        all_sections = (curated_sections + semantic_sections)[:max_cap]

        f03_items_out[item_id] = {
            "item_id": item_id,
            "canonical_topic": topic or None,
            "canonical_description": description or None,
            "canonical_ref_raw": canonical_ref_raw,
            "refs": refs_out,
            "sections": all_sections,
        }
        log.debug(
            "retrieval_map item=%s curated=%d semantic=%d total=%d",
            item_id,
            len(curated_sections),
            len(semantic_sections),
            len(all_sections),
        )

    # ── Process F02 questions（semantic only）────────────────────────────────
    for qid, q in f02_questions.items():
        question_text = q.get("question_text") or ""
        if not question_text:
            continue
        try:
            vectors = embedding_client.embed([question_text])
            hits = store.search(vectors[0], top_k)
        except (EmbeddingError, Exception) as exc:
            log.warning("f02 search failed qid=%s err=%s", qid, type(exc).__name__)
            hits = []

        sections: list[dict] = []
        for hit in hits:
            if score_threshold is not None and (hit.get("score") or 0.0) < score_threshold:
                continue
            sections.append(
                {
                    "reg_code": hit["reg_code"],
                    "section_path": hit["section_path"],
                    "title": hit["title"],
                    "excerpt": hit["text"][:max_excerpt],
                    "score": hit.get("score"),
                    "origin": "semantic",
                }
            )
        f02_questions_out[qid] = {"qid": qid, "sections": sections[:max_cap]}

    if _owns_store:
        store.close()

    # ── Compose + atomic write ───────────────────────────────────────────────
    retrieval_map = {
        "schema_version": 1,
        "built_at": datetime.now(tz=timezone.utc).isoformat(),
        "embedding_model": cfg["embedding_model"],
        "f03_items": f03_items_out,
        "f02_questions": f02_questions_out,
    }

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = output_path.with_name(output_path.name + ".tmp")
    try:
        with tmp.open("w", encoding="utf-8") as fh:
            json.dump(retrieval_map, fh, ensure_ascii=False, indent=2)
        os.replace(tmp, output_path)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise

    log.info(
        "build_retrieval_map written path=%s f03=%d f02=%d",
        output_path,
        len(f03_items_out),
        len(f02_questions_out),
    )


def run_eval(
    *,
    cfg: dict | None = None,
    store=None,
    embedding_client: EmbeddingClient | None = None,
    items: list | None = None,
    output=None,
) -> None:
    """build ④（--eval）：以 20 項 curated 對（21 筆 ref）為 ground truth，
    報告 semantic recall@k + 分數分佈。stdout 報告供人工定 score_threshold。

    AC 不斷言任何具體 threshold 數值（待人工校準後寫入 config）。
    NOTE: 輸出不得貼進會 commit 的檔案。
    """
    from govcheck.rag.refs import parse_regulation_refs, ref_matches_chunk

    if output is None:
        output = sys.stdout

    cfg = cfg or load_rag_config()
    top_k: int = cfg["top_k"]

    _owns_store = store is None
    if store is None:
        store = RegulationStore.from_config(cfg)
        store.open()
    if embedding_client is None:
        embedding_client = EmbeddingClient.from_config(cfg)
    if items is None:
        items = _read_f03_items()

    total_refs = 0
    recalled = 0
    scores_hit: list[float] = []
    scores_miss: list[float] = []

    print(
        f"=== RAG eval  top_k={top_k}  {datetime.now(tz=timezone.utc).isoformat()} ===",
        file=output,
    )

    for item in items:
        topic = item.get("topic") or ""
        description = item.get("description") or ""
        canonical_ref_raw = item.get("canonical_ref_raw")
        item_id = item.get("item_id", "?")

        if not canonical_ref_raw:
            continue
        refs = parse_regulation_refs(canonical_ref_raw)
        if not refs:
            continue

        query_text = f"{topic}:{description}"
        try:
            vectors = embedding_client.embed([query_text])
            hits = store.search(vectors[0], top_k)
        except Exception as exc:
            print(
                f"  WARN: search failed item={item_id} err={type(exc).__name__}",
                file=output,
            )
            continue

        hit_scores: dict[tuple, float] = {
            (h["reg_code"], h["section_path"]): h.get("score", 0.0) for h in hits
        }

        for ref in refs:
            total_refs += 1
            matched_score: float | None = None
            for (hr, hp), sc in hit_scores.items():
                if ref_matches_chunk(ref, hr, hp):
                    matched_score = sc
                    break

            if matched_score is not None:
                recalled += 1
                scores_hit.append(matched_score)
            else:
                if hits:
                    scores_miss.append(min(hit_scores.values()))

    recall = recalled / total_refs if total_refs else 0.0
    print(f"\nrecall@{top_k}: {recalled}/{total_refs} ({recall * 100:.1f}%)", file=output)

    if scores_hit:
        print(
            f"score (hit):  min={min(scores_hit):.3f}  "
            f"max={max(scores_hit):.3f}  "
            f"mean={sum(scores_hit)/len(scores_hit):.3f}",
            file=output,
        )
    if scores_miss:
        print(
            f"score (miss): max_top={max(scores_miss):.3f}",
            file=output,
        )

    # NOTE: 以下僅作人工校準參考，輸出不得貼進會 commit 的檔案
    print(
        "\n(以上報告供人工定 score_threshold；校準後寫入 llm_config.yaml rag.score_threshold)",
        file=output,
    )

    if _owns_store:
        store.close()


# ── CLI main ─────────────────────────────────────────────────────────────────


def main() -> None:
    """argparse CLI：--dry-run ③ / --eval ④ / 預設 ①+②。"""
    load_dotenv()  # repo root .env（若存在）→ os.environ；GOVCHECK_RAG_* 才讀得到覆寫值
    parser = argparse.ArgumentParser(
        description="Build RAG regulation index (①③) + canonical retrieval map (②④)"
    )
    parser.add_argument(
        "--uri",
        help="Milvus URI（覆寫 config；.db = Lite，http:// = Server）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="印 chunk 樹後停止（不 embed 不寫庫；AC-17）",
    )
    parser.add_argument(
        "--eval",
        action="store_true",
        help="跑 recall@k eval 並印報告（需已建索引）",
    )
    args = parser.parse_args()

    cfg = load_rag_config()
    if args.uri:
        cfg = dict(cfg)
        cfg["milvus_uri"] = args.uri

    if args.dry_run:
        print_chunk_tree(cfg=cfg)
        return

    if args.eval:
        store = RegulationStore.from_config(cfg)
        store.open()
        client = EmbeddingClient.from_config(cfg)
        try:
            run_eval(cfg=cfg, store=store, embedding_client=client)
        finally:
            store.close()
        return

    # Full build ①②
    build_index(cfg=cfg)
    build_retrieval_map(cfg=cfg)


if __name__ == "__main__":
    main()
