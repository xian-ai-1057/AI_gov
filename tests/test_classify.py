"""批次自動分類 + 路由 + 編排測試。

分為兩類：
  - 無須官方範本者（空檔/壞檔/PDF/純路由邏輯）→ 恆可跑，提供無網路確定性基線。
  - 須官方範本者（F01/F02/F03 真檔分類與端到端）→ data/original 未連結時 skip。

關鍵驗證：分類依**工作表名稱**而非檔名（刻意給誤導性檔名證明之）。
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from govcheck.classify import (
    FileClassification,
    FileKind,
    classify_file,
    classify_fileobj,
    classify_files,
    route_classifications,
)
from govcheck.parsers.f02_parser import ASSESS_SHEET
from govcheck.review.config import load_review_config
from govcheck.review.engine import review_files, review_routed, review_submission
from tests.fixture_builder import (
    OFFICIAL,
    OFFICIAL_F01,
    OFFICIAL_F03,
    build_f01_fixture,
    build_f02_fixture,
    build_f03_fixture,
)

_CFG = load_review_config()
F01_SHEET = _CFG["f01"]["main_sheet"]
F03_SHEET = _CFG["f03"]["sheet"]

requires_templates = pytest.mark.skipif(
    not (OFFICIAL.exists() and OFFICIAL_F01.exists() and OFFICIAL_F03.exists()),
    reason="官方範本不存在（data/original 未連結）",
)


def codes(report) -> set[str]:
    return {f.code for f in report.findings}


# ── 分類器：無須官方範本 ─────────────────────────────────────────────


def test_pdf_is_supporting(tmp_path):
    p = tmp_path / "R06_可解釋性說明.pdf"
    p.write_bytes(b"%PDF-1.4 not a real pdf")
    c = classify_file(p)
    assert c.kind is FileKind.SUPPORTING
    assert c.matched_sheet is None


def test_docx_is_supporting(tmp_path):
    p = tmp_path / "委外技術文件.docx"
    p.write_bytes(b"PK\x03\x04 junk")  # 不會被開啟（副檔名閘門）
    assert classify_file(p).kind is FileKind.SUPPORTING


def test_blank_xlsx_is_unknown(tmp_path):
    p = tmp_path / "其他試算表.xlsx"
    openpyxl.Workbook().save(p)  # 無任何已知表單分頁
    c = classify_file(p)
    assert c.kind is FileKind.UNKNOWN
    assert c.matched_sheet is None


def test_corrupt_xlsx_is_unknown(tmp_path):
    p = tmp_path / "bad.xlsx"
    p.write_text("this is not an excel file")  # 比照 test_parse_error_does_not_crash
    c = classify_file(p)  # 不應丟例外
    assert c.kind is FileKind.UNKNOWN
    assert "無法開啟" in c.reason


def test_classify_fileobj_matches_extension_gate():
    # 位元組版（UI 預覽用）：非 Excel 檔名 → 佐證，不開檔
    assert classify_fileobj(io.BytesIO(b"x"), "evidence.pdf").kind is FileKind.SUPPORTING


def test_classify_fileobj_blank_xlsx_unknown(tmp_path):
    buf = io.BytesIO()
    openpyxl.Workbook().save(buf)
    buf.seek(0)
    assert classify_fileobj(buf, "x.xlsx").kind is FileKind.UNKNOWN


# ── 路由邏輯：無須官方範本（直接組 FileClassification） ───────────────


def _fc(kind: FileKind, name: str, path: str | None = None) -> FileClassification:
    return FileClassification(path=path or f"/fake/{name}", filename=name, kind=kind, reason="t")


def test_route_summary_always_present():
    _, _, findings = route_classifications([_fc(FileKind.SUPPORTING, "a.pdf")])
    assert findings[0].code == "CLASSIFY.SUMMARY"
    assert findings[0].severity.value == "info"


def test_route_supporting_and_unknown():
    results = [
        _fc(FileKind.F01, "f01.xlsx", "/t/f01.xlsx"),
        _fc(FileKind.SUPPORTING, "R07_公平性.pdf"),
        FileClassification(filename="weird.xlsx", kind=FileKind.UNKNOWN, reason="無已知分頁"),
    ]
    files, supporting, findings = route_classifications(results)
    assert files == {"f01": "/t/f01.xlsx"}
    assert supporting == ["R07_公平性.pdf"]
    cs = {f.code for f in findings}
    assert "CLASSIFY.SUMMARY" in cs
    assert "CLASSIFY.UNRECOGNIZED" in cs


def test_route_duplicate_first_wins():
    results = [
        _fc(FileKind.F02, "first.xlsm", "/t/first.xlsm"),
        _fc(FileKind.F02, "second.xlsm", "/t/second.xlsm"),
    ]
    files, _, findings = route_classifications(results)
    assert files["f02"] == "/t/first.xlsm"  # 第一份勝出
    dup = [f for f in findings if f.code == "CLASSIFY.DUPLICATE_F02"]
    assert len(dup) == 1
    assert dup[0].severity.value == "warn"
    assert "first.xlsm" in dup[0].message and "second.xlsm" in dup[0].message


def test_route_unknown_not_routed():
    results = [FileClassification(filename="x.xlsx", kind=FileKind.UNKNOWN, reason="r")]
    files, supporting, _ = route_classifications(results)
    assert files == {} and supporting == []


# ── 編排 review_files：無須官方範本 ──────────────────────────────────


def test_review_files_returns_tuple(tmp_path):
    p = tmp_path / "x.pdf"
    p.write_bytes(b"%PDF")
    out = review_files([p])
    assert isinstance(out, tuple) and len(out) == 2
    report, results = out
    assert results[0].kind is FileKind.SUPPORTING


def test_review_files_all_supporting_missing_core(tmp_path):
    p = tmp_path / "佐證.pdf"
    p.write_bytes(b"%PDF")
    report, _ = review_files([p])
    c = codes(report)
    assert "DOC.MISSING_F01" in c  # 缺核心表單 → ERROR
    assert "DOC.MISSING_F03" in c
    assert "CLASSIFY.SUMMARY" in c
    assert "SUBMISSION.OK" not in c  # 有 ERROR → 不補 OK


def test_review_files_unknown_excel_warns(tmp_path):
    p = tmp_path / "其他.xlsx"
    openpyxl.Workbook().save(p)
    report, results = review_files([p])
    assert "CLASSIFY.UNRECOGNIZED" in codes(report)
    assert results[0].kind is FileKind.UNKNOWN


def test_review_routed_ok_only_when_clean():
    # 空 files → review_submission 報缺件 ERROR；併入 WARN 後仍不補 OK
    warn = review_submission({}).findings  # 取得一個現成 ERROR 集
    report = review_routed({}, [], class_findings=[])
    assert "SUBMISSION.OK" not in codes(report)
    assert any(f.severity.value == "error" for f in report.findings)
    assert warn  # sanity：空送件本就有缺件 finding


# ── 須官方範本：真檔分類證明「依分頁、非依檔名」+ 端到端 ───────────────


@requires_templates
def test_f01_classified_despite_misleading_name(tmp_path):
    p = build_f01_fixture(tmp_path / "風險評鑑_最終版.xlsx")  # 檔名像 F02
    c = classify_file(p)
    assert c.kind is FileKind.F01
    assert c.matched_sheet == F01_SHEET


@requires_templates
def test_f02_classified_despite_misleading_name(tmp_path):
    from tests.test_f02_rules import BASELINE
    p = build_f02_fixture(tmp_path / "f01_系統表.xlsm", answers=BASELINE, cached_grade="低")
    c = classify_file(p)
    assert c.kind is FileKind.F02
    assert c.matched_sheet == ASSESS_SHEET


@requires_templates
def test_f03_classified_despite_misleading_name(tmp_path):
    p = build_f03_fixture(tmp_path / "資料佐證.xlsx")  # 檔名像佐證
    c = classify_file(p)
    assert c.kind is FileKind.F03
    assert c.matched_sheet == F03_SHEET


@requires_templates
def test_classify_files_preserves_order(tmp_path):
    from tests.test_f02_rules import BASELINE
    paths = [
        build_f03_fixture(tmp_path / "a.xlsx"),
        build_f02_fixture(tmp_path / "b.xlsm", answers=BASELINE, cached_grade="低"),
        build_f01_fixture(tmp_path / "c.xlsx"),
    ]
    kinds = [r.kind for r in classify_files(paths)]
    assert kinds == [FileKind.F03, FileKind.F02, FileKind.F01]


@requires_templates
def test_review_files_mixed_unordered_ok(tmp_path):
    from tests.test_f02_rules import BASELINE
    pdf = tmp_path / "R06.pdf"
    pdf.write_bytes(b"%PDF")
    paths = [  # 亂序 + 佐證夾雜
        build_f03_fixture(tmp_path / "f03.xlsx"),
        pdf,
        build_f02_fixture(tmp_path / "f02.xlsm", answers=BASELINE, cached_grade="低"),
        build_f01_fixture(tmp_path / "f01.xlsx"),
    ]
    report, results = review_files(paths)
    c = codes(report)
    assert "SUBMISSION.OK" in c
    assert "CLASSIFY.SUMMARY" in c
    assert report.error_count == 0 and report.warn_count == 0
    kinds = {r.kind for r in results}
    assert {FileKind.F01, FileKind.F02, FileKind.F03, FileKind.SUPPORTING} <= kinds


@requires_templates
def test_review_files_duplicate_f02_warns(tmp_path):
    from tests.test_f02_rules import BASELINE
    paths = [
        build_f01_fixture(tmp_path / "f01.xlsx"),
        build_f02_fixture(tmp_path / "f02_a.xlsm", answers=BASELINE, cached_grade="低"),
        build_f02_fixture(tmp_path / "f02_b.xlsm", answers=BASELINE, cached_grade="低"),
        build_f03_fixture(tmp_path / "f03.xlsx"),
    ]
    report, _ = review_files(paths)
    c = codes(report)
    assert "CLASSIFY.DUPLICATE_F02" in c
    assert "SUBMISSION.OK" not in c  # 有 WARN → 不補 OK
