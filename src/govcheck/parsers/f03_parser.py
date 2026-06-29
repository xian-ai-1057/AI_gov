"""F03 (.xlsx) 解析器。

- parse_f03_identity：Phase 2 最小識別欄（System Owner，B1）供跨表一致性比對。
- parse_f03_checklist：Phase 3 完整檢核項（含兩段佐證欄 I/J），供規則缺漏檢查與 LLM 判讀。
座標皆來自 review_config.yaml 的 f03 區段。
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path

import openpyxl

from govcheck.models import F03Checklist, F03ChecklistItem, F03Identity
from govcheck.parsers._util import clean
from govcheck.review.config import load_review_config


def parse_f03_identity(path: str | Path, cfg: dict | None = None) -> F03Identity:
    cfg = cfg or load_review_config()
    f03cfg = cfg["f03"]
    path = Path(path)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)

    sheet = f03cfg["sheet"]
    if sheet not in wb.sheetnames:
        wb.close()
        return F03Identity(sheet_present=False)

    owner = clean(wb[sheet][f03cfg["owner_cell"]].value)
    wb.close()
    # 範本預設 B1="XXX"，視同未填，避免誤報跨表 owner 不一致
    if _is_placeholder(owner, f03cfg):
        owner = None
    return F03Identity(system_owner=owner, sheet_present=True)


def _is_placeholder(value: str | None, f03cfg: dict) -> bool:
    if not value:
        return False
    placeholders = {p.casefold() for p in f03cfg.get("placeholder_values", [])}
    return value.casefold() in placeholders


def parse_f03_checklist(path: str | Path, cfg: dict | None = None) -> F03Checklist:
    """解析 F03「檢核表」完整檢核項（含兩段佐證欄 I/J）。

    生命週期欄為合併儲存格，僅群組首列有值 → 往下沿用最後出現值。
    以項次欄（A）符合 item_id_pattern 認定資料列，自動略過空列與備註列。
    """
    cfg = cfg or load_review_config()
    f03cfg = cfg["f03"]
    lc = f03cfg["checklist"]
    path = Path(path)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)

    sheet = f03cfg["sheet"]
    if sheet not in wb.sheetnames:
        wb.close()
        return F03Checklist(sheet_present=False)
    ws = wb[sheet]

    owner = clean(ws[f03cfg["owner_cell"]].value)
    subject = None if _is_placeholder(owner, f03cfg) else owner

    id_pat = re.compile(lc["item_id_pattern"])
    id_col = lc["item_id_col"]
    life_col = lc["lifecycle_col"]
    topic_col = lc["topic_col"]
    desc_col = lc["item_desc_col"]
    prop_col = lc["evidence_proposal_col"]
    golive_col = lc["evidence_golive_col"]
    check_cols: dict[str, str] = lc["check_cols"]  # 標籤 → 欄字母

    items: list[F03ChecklistItem] = []
    last_lifecycle: str | None = None
    for r in range(lc["data_start_row"], ws.max_row + 1):
        life = clean(ws[f"{life_col}{r}"].value)
        if life is not None:
            last_lifecycle = life  # 群組首列更新；合併區其餘列沿用

        item_id = clean(ws[f"{id_col}{r}"].value)
        if not item_id or not id_pat.match(item_id):
            continue  # 空列 / 備註列 / 非資料列

        items.append(F03ChecklistItem(
            item_id=item_id,
            lifecycle=last_lifecycle,
            topic=clean(ws[f"{topic_col}{r}"].value),
            description=clean(ws[f"{desc_col}{r}"].value),
            check_state=_check_state(ws, r, check_cols),
            evidence_proposal=clean(ws[f"{prop_col}{r}"].value),
            evidence_golive=clean(ws[f"{golive_col}{r}"].value),
        ))

    wb.close()
    return F03Checklist(subject=subject, items=items, sheet_present=True)


# 視同「未打勾」的儲存格值（正規化後比對）：空白由 clean() → None 處理；
# 公式快取常出現布林 FALSE 或數字 0，須一併視為未勾，避免誤判為已完成。
_UNMARKED_TOKENS = {"false", "0", "none"}


def _is_marked(value) -> bool:
    s = clean(value)
    return bool(s) and s.casefold() not in _UNMARKED_TOKENS


def _check_state(ws, row: int, check_cols: dict[str, str]) -> str | None:
    """F/G/H 哪一欄有打勾 → 回該欄標籤（是/否/不適用）。

    皆未勾 → None；同時勾超過一欄 → 視為模糊（None），不臆測，交人工判讀，
    避免把資料登錄錯誤誤判成已完成而產生假性缺漏提醒。
    """
    marked = [label for label, col in check_cols.items() if _is_marked(ws[f"{col}{row}"].value)]
    return marked[0] if len(marked) == 1 else None
