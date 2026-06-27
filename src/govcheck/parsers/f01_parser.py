"""F01 (.xlsx) 解析器：官方表單 → F01Form。

主表「AI系統業務與利害關係人」橫向：Row2 欄標籤、Row3 說明、Row4+ 資料列；
任一 A~L 欄非空即視為一筆使用者填寫的 AI 應用，逐列攤平成 F01ApplicationRow。
附屬表（資料/模型/平台）表頭列不同，只收「對應專案/服務名稱」欄非空者——避開模板
自帶示範列（示範列不填對應欄）造成跨表內部一致性誤報。所有座標來自 review_config.yaml。
"""

from __future__ import annotations

import warnings
from pathlib import Path

import openpyxl

from govcheck.models import F01ApplicationRow, F01Form, F01SubRef
from govcheck.parsers._util import clean
from govcheck.review.config import load_review_config


def parse_f01(path: str | Path, cfg: dict | None = None) -> F01Form:
    cfg = cfg or load_review_config()
    f01cfg = cfg["f01"]
    path = Path(path)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)

    rows = _read_rows(wb, f01cfg)
    sub_refs = _read_sub_refs(wb, f01cfg["sub_sheets"], f01cfg.get("placeholder_markers", []))
    wb.close()

    subject = rows[0].project_name if rows else None
    return F01Form(subject=subject, rows=rows, sub_refs=sub_refs)


def _read_rows(wb, f01cfg: dict) -> list[F01ApplicationRow]:
    """主表自 data_start_row 起，任一 A~L 非空即一筆資料列。"""
    ws = wb[f01cfg["main_sheet"]]
    columns: dict[str, str] = f01cfg["columns"]  # 欄字母 → 語意鍵
    start = f01cfg["data_start_row"]
    rows: list[F01ApplicationRow] = []
    for r in range(start, ws.max_row + 1):
        raw = {col: clean(ws[f"{col}{r}"].value) for col in columns}
        if not any(raw.values()):
            continue  # 全空列略過
        kwargs = {sem: raw[col] for col, sem in columns.items()}
        rows.append(F01ApplicationRow(row_index=r, raw=raw, **kwargs))
    return rows


def _read_sub_refs(wb, sub_sheets: list[dict], placeholders: list[str]) -> list[F01SubRef]:
    """各附屬表「對應專案/服務名稱」欄，只收非空且非範本佔位字串者（避開模板示範列）。"""
    refs: list[F01SubRef] = []
    for sub in sub_sheets:
        name = sub["sheet"]
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        col = sub["corr_col"]
        for r in range(sub["data_start_row"], ws.max_row + 1):
            val = clean(ws[f"{col}{r}"].value)
            if val and not _is_placeholder(val, placeholders):
                refs.append(F01SubRef(sheet=name, row_index=r, corr_project_name=val))
    return refs


def _is_placeholder(val: str, markers: list[str]) -> bool:
    """範本佔位/說明字串（如「xxx小幫手(對應前頁填寫)」）視同空白。"""
    low = val.casefold()
    return any(m.casefold() in low for m in markers)
