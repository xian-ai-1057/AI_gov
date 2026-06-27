"""從官方 F02 範本程式化生成測試 fixture（正例/反例）。

做法：複製官方 .xlsm，寫入指定的 C 欄答案、加成因子、以及「Excel 已存的」快取分數
（E46:H46 / I2 / I4），再選擇性地在續填表寫入資料列。寫入快取格時用字面值覆蓋公式，
parser 以 data_only 讀回即為我們設定的值——藉此完全掌控每條規則的正/反路徑。

絕不寫回 data/original：輸出一律寫到呼叫端指定的暫存/fixtures 路徑。
"""

from __future__ import annotations

import warnings
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OFFICIAL = ROOT / "data" / "original" / "附件二：AI-R02-F02 AI風險評鑑.xlsm"
OFFICIAL_F01 = ROOT / "data" / "original" / "附件一：AI-R02-F01 AI系統資訊表.xlsx"
OFFICIAL_F03 = ROOT / "data" / "original" / "附件三：AI-R02-F03 AI系統提案規劃暨上線檢核表.xlsx"

ASSESS_SHEET = "AI系統固有風險分級評估表"
RESIDUAL_SHEET = "AI系統剩餘風險評鑑表"
TREATMENT_SHEET = "AI系統風險處理計畫表"

PCT_CELLS = {"finance": "E46", "operation": "F46", "reputation": "G46", "compliance": "H46"}

# --- F01 ---
F01_MAIN_SHEET = "AI系統業務與利害關係人"
F01_DATA_START = 4
# 附屬表對應欄座標（須與 config/review_config.yaml 的 sub_sheets 對齊）
F01_SUB_COORD = {"資料": ("AB", 4), "模型": ("R", 4), "平台": ("I", 3)}
# 預設一筆 A~I 皆填的合規列（H/I 與 F03 預設 owner 對齊，便於跨表正例）
F01_DEFAULT_ROW = {
    "A": "數位金融處/AI應用部", "B": "王小明", "C": "2026.06.27",
    "D": "智能客服小幫手", "E": "1.直接使用AI", "F": "網路銀行",
    "G": "EB001", "H": "李大華", "I": "陳美麗",
    "J": "對話式客服", "K": "自動回覆客戶常見問題", "L": "生成式",
}

# --- F03 ---
F03_SHEET = "檢核表"


def build_f02_fixture(
    out_path: str | Path,
    answers: dict[str, str] | None = None,
    uplift: dict[str, str] | None = None,
    cached_pct: dict[str, float] | None = None,
    cached_overall: float | None = None,
    cached_grade: str | None = None,
    residual_rows: list[list] | None = None,
    treatment_rows: list[list] | None = None,
    filing_unit: str | None = None,
) -> Path:
    out_path = Path(out_path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(OFFICIAL, keep_vba=True)
    ws = wb[ASSESS_SHEET]

    qid_row = _qid_to_row(ws)
    uplift_row = {"factor1": 42, "factor2": 43, "factor3": 44}

    for qid, val in (answers or {}).items():
        if qid not in qid_row:
            raise KeyError(f"未知題號 {qid}")
        ws[f"C{qid_row[qid]}"] = val
    for name, val in (uplift or {}).items():
        ws[f"C{uplift_row[name]}"] = val

    # 覆蓋快取分數（字面值取代公式）
    for d, val in (cached_pct or {}).items():
        ws[PCT_CELLS[d]] = val
    if cached_overall is not None:
        ws["I2"] = cached_overall
    if cached_grade is not None:
        ws["I4"] = cached_grade

    _write_rows(wb[RESIDUAL_SHEET], residual_rows)
    _write_rows(wb[TREATMENT_SHEET], treatment_rows)

    # 風險處理計畫表填表單位（N2）；預設 None → 不寫，行為與既有測試完全一致
    if filing_unit is not None:
        wb[TREATMENT_SHEET]["N2"] = filing_unit

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _qid_to_row(ws) -> dict[str, int]:
    mapping = {}
    for row in ws.iter_rows():
        v = row[0].value
        if isinstance(v, str) and v.strip():
            mapping[v.strip()] = row[0].row
    return mapping


def _write_rows(ws, rows: list[list] | None) -> None:
    if not rows:
        return
    for r, values in enumerate(rows, start=2):  # 表頭在第 1 列
        for c, val in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=val)


def build_f01_fixture(
    out_path: str | Path,
    rows: list[dict[str, str | None]] | None = None,
    sub_refs: dict[str, list[str | None]] | None = None,
) -> Path:
    """從官方 F01 範本生成測試檔。

    rows：每筆是 {欄字母: 值}，自第 4 列起寫主表；預設給一筆 A~I 皆填的合規列。
    sub_refs：{sheet 名: [對應欄各列值]}，None 的列不寫，用來測 F01 內部一致性。
    絕不寫回 data/original：輸出一律寫到呼叫端指定的暫存路徑。
    """
    out_path = Path(out_path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(OFFICIAL_F01)
    ws = wb[F01_MAIN_SHEET]

    if rows is None:
        rows = [dict(F01_DEFAULT_ROW)]
    for i, row in enumerate(rows):
        r = F01_DATA_START + i
        for col, val in row.items():
            ws[f"{col}{r}"] = val

    for sheet_name, values in (sub_refs or {}).items():
        sub_ws = wb[sheet_name]
        col, start = F01_SUB_COORD[sheet_name]
        for j, val in enumerate(values):
            if val is not None:
                sub_ws[f"{col}{start + j}"] = val

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_f03_fixture(out_path: str | Path, system_owner: str | None = "李大華") -> Path:
    """從官方 F03 範本生成測試檔；只寫識別欄 System Owner（B1）。

    預設值與 F01_DEFAULT_ROW 的 H 欄一致，便於跨表 owner 正例。
    """
    out_path = Path(out_path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(OFFICIAL_F03)
    ws = wb[F03_SHEET]
    if system_owner is not None:
        ws["B1"] = system_owner
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
