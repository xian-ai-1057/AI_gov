"""parse_f03_checklist 的單元測試：不需官方範本，可在乾淨環境執行。

涵蓋 _check_state 對 False/0/空白與多重勾選的處理，以及缺工作表時的降級行為。
"""

from __future__ import annotations

import openpyxl

from govcheck.parsers.f03_parser import _check_state, parse_f03_checklist

_COLS = {"是": "F", "否": "G", "不適用": "H"}


def _ws_with(values: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    for cell, val in values.items():
        ws[cell] = val
    return ws


def test_check_state_ignores_false_zero_blank():
    # 公式快取常見的布林 FALSE / 數字 0 應視為未勾，真正的勾在 H
    ws = _ws_with({"F5": False, "G5": 0, "H5": "V"})
    assert _check_state(ws, 5, _COLS) == "不適用"


def test_check_state_single_mark():
    ws = _ws_with({"F5": "V"})
    assert _check_state(ws, 5, _COLS) == "是"


def test_check_state_all_empty_is_none():
    ws = _ws_with({})
    assert _check_state(ws, 5, _COLS) is None


def test_check_state_multi_mark_is_ambiguous_none():
    # 同時勾兩欄（登錄錯誤）→ 不臆測，回 None，避免假性「已完成」
    ws = _ws_with({"F5": "V", "H5": "x"})
    assert _check_state(ws, 5, _COLS) is None


def test_parse_checklist_absent_sheet(tmp_path):
    p = tmp_path / "nosheet.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "其他工作表"
    wb.save(p)
    cl = parse_f03_checklist(p)
    assert cl.sheet_present is False
    assert cl.items == []
