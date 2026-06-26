"""建置第一步：從官方 F02 (.xlsm) 抽出「固有風險分級評估表」的計分常數，產出 f02_scoring.yaml。

做法（不靠人工抄寫，直接解析 Excel 公式，可重現）：
  - 讀「AI系統固有風險分級評估表」E:H 欄每格公式，形如
        =IF($C{row}="Y", 固有風險分數參照表!T${refrow}, 0)
    解析出：該題的計分觸發答案 (Y/N) 與參照列 refrow。
  - 從 A 欄取題號，B 欄取題目，到「固有風險分數參照表」的 T/U/V/W 取四個風險域分數。
  - 加成因子 (rows 42-44) 與正規化分母 / 分級門檻一併寫出。

唯讀來源，本腳本只讀不寫 data/original。輸出寫到 src/govcheck/config/f02_scoring.yaml。

用法：python scripts/extract_f02_scoring.py
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path

import openpyxl
import yaml

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
SRC_XLSM = ROOT / "data" / "original" / "附件二：AI-R02-F02 AI風險評鑑.xlsm"
OUT_YAML = ROOT / "src" / "govcheck" / "config" / "f02_scoring.yaml"

ASSESS_SHEET = "AI系統固有風險分級評估表"
REF_SHEET = "固有風險分數參照表"

# 風險域：對照 E/F/G/H 欄與參照表 T/U/V/W 欄
DOMAINS = ["finance", "operation", "reputation", "compliance"]  # 財務/作業/商譽/合規
REF_COLS = ["T", "U", "V", "W"]

# 題組定義（題號 → 子項數）。單選=恰一個 Y；複選=至少一個 Y；一般=單一 Y/N。
SINGLE_CHOICE = {"UC-04": 4, "UC-05": 4, "UC-06": 3, "D-01": 4, "M-03": 2}
MULTI_CHOICE = {"D-02": 4}
GENERAL = ["UC-01", "UC-02", "UC-03", "UC-07", "D-03", "D-04", "M-01", "M-02"]

# 加成因子：評估表的列 → 觸發答案 Y 時取參照表的哪一格（否則 ×1）
UPLIFT_CELLS = {
    "factor1": 42,
    "factor2": 43,
    "factor3": 44,
}

FORMULA_RE = re.compile(r'IF\(\$C\d+="([YN])",\s*固有風險分數參照表!T\$(\d+)')


def main() -> None:
    # 評估表用 data_only=False 取公式；參照表用 data_only=True 取快取常數（總和列為公式）
    wb = openpyxl.load_workbook(SRC_XLSM, data_only=False)
    wb_val = openpyxl.load_workbook(SRC_XLSM, data_only=True)
    ws = wb[ASSESS_SHEET]
    ref = wb_val[REF_SHEET]

    def ref_scores(refrow: int) -> dict[str, float]:
        return {
            d: _num(ref[f"{c}{refrow}"].value) for d, c in zip(DOMAINS, REF_COLS)
        }

    questions: dict[str, dict] = {}
    for row in range(2, 47):
        qid = ws[f"A{row}"].value
        formula = ws[f"E{row}"].value
        if not isinstance(qid, str) or not isinstance(formula, str):
            continue
        qid = qid.strip()
        m = FORMULA_RE.search(formula)
        if not m:
            continue  # 標題列 / 小計列 / 加成因子另外處理
        trigger, refrow = m.group(1), int(m.group(2))
        questions[qid] = {
            "row": row,
            "score_on": trigger,  # 答此值時計分
            "scores": ref_scores(refrow),
            "text": (ws[f"B{row}"].value or "").strip(),
        }

    # 加成因子：=IF($C{r}="Y", 參照表!$T${refrow}, 1)
    uplift = {}
    uf_re = re.compile(r'固有風險分數參照表!\$T\$(\d+)')
    for name, r in UPLIFT_CELLS.items():
        f = ws[f"E{r}"].value
        refrow = int(uf_re.search(f).group(1))
        uplift[name] = {
            "row": r,
            "multiplier_on_yes": _num(ref[f"T{refrow}"].value),  # 答「有」(Y) 的乘數
            "multiplier_default": 1,
        }

    # 正規化分母（參照表「總和」列 T66:W66）與評估表 row46 公式一致
    denom = {d: _num(ref[f"{c}66"].value) for d, c in zip(DOMAINS, REF_COLS)}

    config = {
        "_source": "附件二：AI-R02-F02 AI風險評鑑.xlsm（由 scripts/extract_f02_scoring.py 自動產出，勿手改）",
        "domains": DOMAINS,
        "domain_labels": {"finance": "財務", "operation": "作業", "reputation": "商譽", "compliance": "合規"},
        "grade_thresholds": {"low_below": 50, "mid_below": 75},  # <50 低, <75 中, 否則 高
        "normalization_denominators": denom,
        "groups": {
            "single_choice": SINGLE_CHOICE,
            "multi_choice": MULTI_CHOICE,
            "general": GENERAL,
        },
        "uplift_factors": uplift,
        "questions": {k: {"score_on": v["score_on"], "scores": v["scores"]}
                      for k, v in questions.items()},
    }

    OUT_YAML.parent.mkdir(parents=True, exist_ok=True)
    with OUT_YAML.open("w", encoding="utf-8") as fh:
        yaml.safe_dump(config, fh, allow_unicode=True, sort_keys=False)
    print(f"✓ wrote {OUT_YAML.relative_to(ROOT)}  ({len(questions)} questions)")


def _num(v) -> float:
    if v is None:
        return 0.0
    f = float(v)
    return int(f) if f == int(f) else f


if __name__ == "__main__":
    main()
